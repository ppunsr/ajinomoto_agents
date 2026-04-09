import openpyxl
import sys

def sort_sheet(file_path, sheet_name, sort_col, start_row, reverse=True):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found")
        return
    
    ws = wb[sheet_name]
    data = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Read data
    for r in range(start_row, max_row + 1):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if any(v is not None for v in row_data):
            data.append(row_data)
        else:
            break
    
    # Sort data
    # sort_col is 1-indexed
    data.sort(key=lambda x: x[sort_col-1] if x[sort_col-1] is not None else 0, reverse=reverse)
    
    # Write back
    for i, row_data in enumerate(data):
        for j, val in enumerate(row_data):
            ws.cell(row=start_row + i, column=j + 1).value = val
            
    wb.save(file_path)
    print(f"Sorted {sheet_name} in {file_path}")

if __name__ == "__main__":
    path = sys.argv[1]
    sort_sheet(path, "menu", 2, 2)
    sort_sheet(path, "score", 3, 2)
