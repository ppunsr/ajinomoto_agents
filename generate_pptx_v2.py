import zipfile
import re
import os
import shutil
import openpyxl
from datetime import datetime

def build_cache(data, is_num=True):
    tag = 'num' if is_num else 'str'
    xml = f'<c:{tag}Cache>'
    if is_num:
        xml += '<c:formatCode>General</c:formatCode>'
    xml += f'<c:ptCount val="{len(data)}"/>'
    for i, v in enumerate(data):
        val = v if v is not None else (0 if is_num else "")
        xml += f'<c:pt idx="{i}"><c:v>{val}</c:v></c:pt>'
    xml += f'</c:{tag}Cache>'
    return xml

def update_pptx(excel_path, template_path, output_path, month):
    print(f"Starting surgical update: {template_path} -> {output_path}")
    
    # 1. Extract data from Excel
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # User Engagement data (Feb + Mar)
    eng_rows = []
    if 'User Engagement' in wb.sheetnames:
        ws = wb['User Engagement']
        for r in range(2, ws.max_row + 1):
            d = ws.cell(row=r, column=1).value
            if isinstance(d, datetime) and d.strftime('%b').lower() in ['feb', 'mar']:
                eng_rows.append((d.strftime('%d/%m'), ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value))

    # 2. Unzip PPTX
    temp_dir = 'temp_pptx_surgical'
    if os.path.exists(temp_dir): shutil.rmtree(temp_dir)
    with zipfile.ZipFile(template_path, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)

    # 3. Update Chart XMLs
    charts_dir = os.path.join(temp_dir, 'ppt', 'charts')
    if os.path.exists(charts_dir):
        for filename in os.listdir(charts_dir):
            if not filename.endswith('.xml'): continue
            filepath = os.path.join(charts_dir, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            orig_content = content
            
            # Remove old caches
            content = re.sub(r'<c:(num|str)Cache>.*?</c:\1Cache>', '', content, flags=re.DOTALL)
            
            if filename in ['chart1.xml', 'chart2.xml'] and eng_rows:
                # Update formulas
                content = content.replace("'User Engagement (2)'!", "'User Engagement'!")
                content = content.replace("$A$67:$A$123", "$A$36:$A$125")
                content = content.replace("$B$67:$B$123", "$B$36:$B$125")
                content = content.replace("$C$67:$C$123", "$C$36:$C$125")
                
                # Inject Categories Cache (Dates)
                cat_cache = build_cache([r[0] for r in eng_rows], is_num=False)
                content = re.sub(r'(<c:cat>.*?<c:(num|str)Ref>.*?<c:f>.*?</c:f>)', r'\1' + cat_cache, content, flags=re.DOTALL)
                
                # Inject Series 0 Cache (New)
                val0_cache = build_cache([r[1] for r in eng_rows], is_num=True)
                content = re.sub(r'(<c:ser>.*?<c:idx val="0".*?<c:val>.*?<c:numRef>.*?<c:f>.*?</c:f>)', r'\1' + val0_cache, content, flags=re.DOTALL)
                
                # Inject Series 1 Cache (Returning)
                val1_cache = build_cache([r[2] for r in eng_rows], is_num=True)
                content = re.sub(r'(<c:ser>.*?<c:idx val="1".*?<c:val>.*?<c:numRef>.*?<c:f>.*?</c:f>)', r'\1' + val1_cache, content, flags=re.DOTALL)

                # Hardcode Labels
                content = re.sub(r'<c:strRef>\s*<c:f>.*?\$B\$1</c:f>.*?</c:strRef>', '<c:v>New user</c:v>', content)
                content = re.sub(r'<c:strRef>\s*<c:f>.*?\$C\$1</c:f>.*?</c:strRef>', '<c:v>returning User</c:v>', content)

            # Save if changed
            if content != orig_content:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(content)
                print(f"Surgically updated {filename}")

    # 4. Update Slides (Text)
    slides_dir = os.path.join(temp_dir, 'ppt', 'slides')
    for filename in os.listdir(slides_dir):
        if not filename.endswith('.xml') or '_' in filename: continue
        filepath = os.path.join(slides_dir, filename)
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        orig_content = content
        if month.lower().startswith('mar'):
            content = content.replace("February-2026", "March-2026")
        content = content.replace("2025/11/28 – 2026/03/31", "2025/11/28 - 2026/04/07")
        
        if content != orig_content:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)

    # 5. Zip back
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                fpath = os.path.join(root, file)
                arcname = os.path.relpath(fpath, temp_dir)
                zipf.write(fpath, arcname)
    
    shutil.rmtree(temp_dir)
    print(f"Successfully generated {output_path}")

if __name__ == "__main__":
    update_pptx("Aji_game copy_March.xlsx", "Merkle Thailand -Ajipanda's Kitchen report- 260331  copy.pptx", "Report_March.pptx", "March")
