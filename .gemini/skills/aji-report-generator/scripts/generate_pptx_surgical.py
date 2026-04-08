
import zipfile
import re
import sys
import os
import shutil
import openpyxl
from datetime import datetime

def build_num_cache(data, format_code='General'):
    xml = f'<c:numCache><c:formatCode>{format_code}</c:formatCode><c:ptCount val="{len(data)}"/>'
    for i, val in enumerate(data):
        v = val if val is not None else 0
        xml += f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>'
    xml += '</c:numCache>'
    return xml

def build_str_cache(data):
    xml = f'<c:strCache><c:ptCount val="{len(data)}"/>'
    for i, val in enumerate(data):
        v = str(val) if val is not None else ''
        xml += f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>'
    xml += '</c:strCache>'
    return xml

def update_pptx(excel_path, template_path, output_path, month):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    temp_dir = 'temp_pptx_unzip'
    if os.path.exists(temp_dir): shutil.rmtree(temp_dir)
    with zipfile.ZipFile(template_path, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)

    # Simplified data fetching
    eng_rows = []
    ws_ue = wb['User Engagement']
    for r in range(2, ws_ue.max_row + 1):
        d = ws_ue.cell(row=r, column=1).value
        if isinstance(d, datetime) and d.strftime('%b').lower() in ['feb', 'mar']:
            eng_rows.append((d.strftime('%d/%m'), ws_ue.cell(row=r, column=2).value, ws_ue.cell(row=r, column=3).value))
    
    score_rows = []
    ws_sc = wb['gameplay_report(score) ']
    for r in range(2, ws_sc.max_row + 1):
        d = ws_sc.cell(row=r, column=1).value
        if isinstance(d, datetime) and d.strftime('%b').lower() == month.lower()[:3]:
            score_rows.append((d.strftime('%d/%m'), ws_sc.cell(row=r, column=2).value))

    # Surgical replacement without complex backreferences in string literals
    charts_dir = os.path.join(temp_dir, 'ppt', 'charts')
    for filename in os.listdir(charts_dir):
        if not filename.endswith('.xml'): continue
        path = os.path.join(charts_dir, filename)
        with open(path, 'r', encoding='utf-8') as f: content = f.read()
        
        orig = content
        content = re.sub(r'<c:numCache>.*?</c:numCache>', '', content, flags=re.DOTALL)
        content = re.sub(r'<c:strCache>.*?</c:strCache>', '', content, flags=re.DOTALL)
        
        if filename == 'chart1.xml' or filename == 'chart2.xml':
            if eng_rows:
                # Use string find/replace for formula to avoid escaping hell
                content = content.replace(''User Engagement (2)'!7:23', ''User Engagement'!6:25')
                # Append cache after formula tag manually
                content = content.replace('</c:f>', '</c:f>' + build_num_cache([0]*len(eng_rows)), 1) 
        
        if content != orig:
            with open(path, 'w', encoding='utf-8') as f: f.write(content)

    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                fpath = os.path.join(root, file)
                arcname = os.path.relpath(fpath, temp_dir)
                zipf.write(fpath, arcname)
    shutil.rmtree(temp_dir)
    print(f'Saved to {output_path}')

if __name__ == '__main__':
    update_pptx(sys.argv[1], sys.argv[2], sys.argv[4], sys.argv[3])
