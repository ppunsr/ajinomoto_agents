import zipfile
import re
import sys
import os
import shutil
import openpyxl
import datetime

def get_target_months(target_month_str, include_previous=False):
    target_month_str = target_month_str.lower()[:3]
    try:
        dt = datetime.datetime.strptime(target_month_str, "%b")
        months = [dt.strftime("%b").lower(), dt.strftime("%B").lower()]
        
        if include_previous:
            prev_month = dt.month - 1 if dt.month > 1 else 12
            dt_prev = datetime.datetime(2000, prev_month, 1)
            months.extend([dt_prev.strftime("%b").lower(), dt_prev.strftime("%B").lower()])
        return set(months)
    except:
        return {target_month_str}

def get_month_location(ws, target_month_str, include_previous=False):
    target_months = get_target_months(target_month_str, include_previous)
    
    start_row = None
    end_row = None
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        is_match = False
        if isinstance(cell_val, datetime.datetime):
            if cell_val.strftime('%b').lower() in target_months or cell_val.strftime('%B').lower() in target_months:
                is_match = True
        elif isinstance(cell_val, str):
            if any(m in cell_val.lower() for m in target_months):
                is_match = True
                
        if is_match:
            if start_row is None:
                start_row = row
            end_row = row
            
    if start_row and end_row:
        return ('row', start_row, end_row)
        
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell_val = cell.value
            is_match = False
            if isinstance(cell_val, datetime.datetime):
                if cell_val.strftime('%b').lower() in target_months or cell_val.strftime('%B').lower() in target_months:
                    is_match = True
            elif isinstance(cell_val, str):
                if any(m in cell_val.lower() for m in target_months) and len(cell_val.strip()) < 15:
                    is_match = True
            if is_match:
                if not include_previous or cell_val.lower().startswith(target_month_str.lower()[:3]):
                    return ('col', cell.column_letter)
                
    return None

def sort_static_sheets(wb):
    sort_configs = {
        'menu': {'sort_col': 2, 'start_row': 2},
        'score': {'sort_col': 3, 'start_row': 2}
    }
    for sheet_name, cfg in sort_configs.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            start_row = cfg['start_row']
            sort_col = cfg['sort_col']
            data = []
            for r in range(start_row, ws.max_row + 1):
                row_data = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in row_data):
                    data.append(row_data)
                else:
                    break
            if data:
                data.sort(key=lambda x: x[sort_col-1] if x[sort_col-1] is not None else 0, reverse=False)
                for i, row_data in enumerate(data):
                    for j, val in enumerate(row_data):
                        ws.cell(row=start_row + i, column=j + 1).value = val
                print(f"Sorted sheet '{sheet_name}' ASC by column {sort_col}")

def update_graphs_zip(input_file, output_file, target_month):
    print(f"Reading target locations for '{target_month}' from {input_file}...")
    try:
        wb = openpyxl.load_workbook(input_file)
    except Exception as e:
        print(f"Failed to open {input_file}: {e}")
        return

    # Sort the static sheets and save to a temporary file
    sort_static_sheets(wb)
    temp_excel = "temp_sorted.xlsx"
    wb.save(temp_excel)
    
    # Reload data-only to get correct rows
    wb_data = openpyxl.load_workbook(temp_excel, data_only=True)

    sheets_of_interest = ['User_funnel', 'User Engagement', 'gameplay_report(score) ', 'gameplay_report(time) ']
    
    transforms = {}
    for sheet_name in sheets_of_interest:
        if sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]
            include_prev = (sheet_name in ['User Engagement', 'gameplay_report(time) ', 'gameplay_report(score) '])
            loc = get_month_location(ws, target_month, include_previous=include_prev)
            if loc:
                transforms[sheet_name] = loc
            else:
                print(f"Sheet '{sheet_name}': Could not find data.")
                
    static_sheets = {
        'state': {'start': 12, 'col': 1},
        'menu': {'start': 2, 'col': 1},
        'score': {'start': 2, 'col': 2}
    }
    
    for sheet_name, cfg in static_sheets.items():
        if sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]
            start_row = cfg['start']
            col = cfg['col']
            end_row = start_row - 1
            
            for row in range(start_row, ws.max_row + 2):
                val = ws.cell(row=row, column=col).value
                if val is None or str(val).strip() == "":
                    break
                end_row = row
                
            if end_row >= start_row:
                transforms[sheet_name] = ('row', start_row, end_row)
                print(f"Sheet '{sheet_name}': Automatically bound to rows {start_row} to {end_row}.")
    
    wb_data.close()
    
    if not transforms:
        print("No ranges found to update. Exiting.")
        os.remove(temp_excel)
        return

    temp_dir = 'temp_excel_unzip'
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
        
    print(f"Unzipping {input_file} to modify chart XML directly (preserves all unsupported objects)...")
    with zipfile.ZipFile(input_file, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
        
    # Inject sorted sheetData from temp_excel into original XMLs
    with zipfile.ZipFile(temp_excel, 'r') as z_sorted:
        for sheet_idx, xml_path in [('6', 'xl/worksheets/sheet6.xml'), ('7', 'xl/worksheets/sheet7.xml')]:
            try:
                sorted_xml = z_sorted.read(xml_path).decode('utf-8')
                sheet_data_match = re.search(r'<sheetData>.*?</sheetData>', sorted_xml, flags=re.DOTALL)
                
                if sheet_data_match:
                    orig_path = os.path.join(temp_dir, xml_path)
                    with open(orig_path, 'r', encoding='utf-8') as f:
                        orig_xml = f.read()
                        
                    patched_xml = re.sub(r'<sheetData>.*?</sheetData>', sheet_data_match.group(0), orig_xml, flags=re.DOTALL)
                    with open(orig_path, 'w', encoding='utf-8') as f:
                        f.write(patched_xml)
                    print(f"Injected sorted data into {xml_path}")
            except Exception as e:
                print(f"Failed to inject sorted data for {xml_path}: {e}")

    os.remove(temp_excel)

    updated_files = 0
    
    def process_formula(full_ref):
        if '!' in full_ref:
            sheet_part, cell_part = full_ref.rsplit('!', 1)
            sheet_name_clean = sheet_part.strip("'")
            
            target_transform = None
            if sheet_name_clean in transforms:
                target_transform = transforms[sheet_name_clean]
            else:
                for k, v in transforms.items():
                    if k.strip() == sheet_name_clean.strip():
                        target_transform = v
                        break
                        
            if target_transform:
                if target_transform[0] == 'row':
                    new_start, new_end = target_transform[1], target_transform[2]
                    if ':' in cell_part:
                        left, right = cell_part.split(':')
                        left = re.sub(r'\d+', str(new_start), left)
                        right = re.sub(r'\d+', str(new_end), right)
                        new_cell_part = f"{left}:{right}"
                    else:
                        if not re.search(r'\$1$', cell_part) and not re.search(r'[A-Z]1$', cell_part):
                            new_cell_part = re.sub(r'\d+', str(new_start), cell_part)
                        else:
                            new_cell_part = cell_part
                else: # 'col'
                    new_col = target_transform[1]
                    def repl_col(part):
                        col_match = re.search(r'[A-Z]+', part)
                        if col_match and col_match.group(0) == 'A':
                            return part 
                        return re.sub(r'[A-Z]+', new_col, part)
                    
                    if ':' in cell_part:
                        left, right = cell_part.split(':')
                        new_cell_part = f"{repl_col(left)}:{repl_col(right)}"
                    else:
                        new_cell_part = repl_col(cell_part)
                        
                return f"{sheet_part}!{new_cell_part}"
        return full_ref

    # 1. Update Charts XML
    charts_dir = os.path.join(temp_dir, 'xl', 'charts')
    if os.path.exists(charts_dir):
        for filename in os.listdir(charts_dir):
            if filename.endswith('.xml'):
                filepath = os.path.join(charts_dir, filename)
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                original_content = content
                
                def replace_c_f(match):
                    tag = match.group(1)
                    full_ref = match.group(2)
                    new_ref = process_formula(full_ref)
                    return f"<{tag}:f>{new_ref}</{tag}:f>"

                content = re.sub(r'<(c|cx):f>(.*?)</\1:f>', replace_c_f, content)
                
                def preserve_format(match):
                    tag = match.group(1)
                    format_code = re.search(f'<{tag}:formatCode>.*?</{tag}:formatCode>', match.group(0))
                    if format_code:
                        return f'<{tag}:numCache>{format_code.group(0)}</{tag}:numCache>'
                    return ''
                
                content = re.sub(r'<(c|cx):numCache>.*?</\1:numCache>', preserve_format, content, flags=re.DOTALL)
                content = re.sub(r'<(c|cx):strCache>.*?</\1:strCache>', '', content, flags=re.DOTALL)
                
                # Hardcode names for User Engagement legend
                if "new_users" in content or "returning_users" in content or "'User Engagement'!$B$1" in content:
                    content = re.sub(r'<c:strRef>\s*<c:f>\'User Engagement\'!\$B\$1</c:f>.*?</c:strRef>', '<c:v>New user</c:v>', content)
                    content = re.sub(r'<c:strRef>\s*<c:f>\'User Engagement\'!\$C\$1</c:f>.*?</c:strRef>', '<c:v>returning User</c:v>', content)
                
                # Apply gameplay_report(time) formatting to gameplay_report(score) labels (red box, white text)
                if "'gameplay_report(score) '" in content:
                    content = content.replace('<a:schemeClr val="lt1"/>', '<a:srgbClr val="C00000"/>')
                    content = re.sub(r'(<a:ln[^>]*>.*?<a:solidFill>\s*)<a:srgbClr val="C00000"/>(\s*</a:solidFill>)', r'\g<1><a:schemeClr val="accent3"/>\g<2>', content)
                    content = content.replace('<a:schemeClr val="dk1"/>', '<a:schemeClr val="bg1"/>')
                
                if content != original_content:
                    with open(filepath, 'w', encoding='utf-8') as f:
                        f.write(content)
                    print(f"Updated references in {filename}")
                    updated_files += 1

    # 2. Update Workbook XML (for named ranges used by Funnel charts etc.)
    workbook_path = os.path.join(temp_dir, 'xl', 'workbook.xml')
    if os.path.exists(workbook_path):
        with open(workbook_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        original_content = content
        
        def replace_defined_name(match):
            prefix = match.group(1)
            full_ref = match.group(2)
            suffix = match.group(3)
            new_ref = process_formula(full_ref)
            return f"{prefix}{new_ref}{suffix}"

        content = re.sub(r'(<definedName[^>]*>)(.*?)(</definedName>)', replace_defined_name, content)
        
        if content != original_content:
            with open(workbook_path, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f"Updated named ranges in workbook.xml")
            updated_files += 1

    if updated_files > 0:
        print(f"Zipping modified contents into {output_file}...")
        with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zipf.write(file_path, arcname)
        print(f"Successfully created {output_file}")
    else:
        print("No chart references matched the targeted sheets. No new file created.")

    shutil.rmtree(temp_dir)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python update_graphs.py <input_file> <target_month> [output_file]")
        sys.exit(1)
        
    input_file = sys.argv[1]
    target_month = sys.argv[2]
    
    if len(sys.argv) > 3:
        output_file = sys.argv[3]
    else:
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_{target_month}{ext}"
        
    update_graphs_zip(input_file, output_file, target_month)
