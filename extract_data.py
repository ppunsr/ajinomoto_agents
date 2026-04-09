import openpyxl
import json
import sys
import calendar
from datetime import datetime, date

def extract_data(excel_path, month_str):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    m_target = month_str.lower()[:3]
    m_prev = 'feb' if m_target == 'mar' else 'jan'
    year = 2026 
    month_num = list(calendar.month_abbr).index(month_str[:3].capitalize())
    prev_month_num = list(calendar.month_abbr).index(m_prev.capitalize())
    
    target_full = calendar.month_name[month_num]
    prev_full = calendar.month_name[prev_month_num]
    
    last_day = calendar.monthrange(year, month_num)[1]
    report_end_date = date(year, month_num, last_day)

    res_stats = {'dau': 0, 'prev_dau': 0, 'mau': 0, 'prev_mau': 0, 'stickiness': 0, 'prev_stickiness': 0, 'avg_score': 0, 'score_change': 0, 'avg_time': 0, 'time_change': 0}
    
    # 1. User Funnel
    ws_f = wb['User_funnel']
    f_col = 3 if m_target == 'mar' else 2
    click = ws_f.cell(row=2, column=f_col).value or 0
    reg = ws_f.cell(row=3, column=f_col).value or 0
    play = ws_f.cell(row=4, column=f_col).value or 0
    
    conv_reg = (reg/click*100) if click else 0
    drop_off = ((reg-play)/reg*100) if reg else 0

    # 2. User Engagement
    ws_ue = wb['User Engagement']
    for r in range(1, 15):
        for c in range(1, 15):
            v = ws_ue.cell(row=r, column=c).value
            if isinstance(v, str):
                if v == 'Monthly active user': 
                    res_stats['mau'] = ws_ue.cell(row=r, column=c + (4 if m_target == 'mar' else 3)).value or 0
                    res_stats['prev_mau'] = ws_ue.cell(row=r, column=c + (3 if m_target == 'mar' else 2)).value or 0
                if v == 'user stickiness': 
                    res_stats['stickiness'] = ws_ue.cell(row=r, column=c + (4 if m_target == 'mar' else 3)).value or 0
                    res_stats['prev_stickiness'] = ws_ue.cell(row=r, column=c + (3 if m_target == 'mar' else 2)).value or 0
                if 'Daily active' in v or 'Daily actuve' in v:
                    res_stats['dau'] = ws_ue.cell(row=r, column=c + (4 if m_target == 'mar' else 3)).value or 0
                    res_stats['prev_dau'] = ws_ue.cell(row=r, column=c + (3 if m_target == 'mar' else 2)).value or 0

    # 3. Score
    ws_sc = wb['gameplay_report(score) ']
    for c in range(1, 10):
        v = ws_sc.cell(row=1, column=c).value
        if v == f'AVG({month_str.capitalize()})': 
            res_stats['avg_score'] = ws_sc.cell(row=2, column=c).value or 0
    prev_sc_label = f'AVG({m_prev.capitalize()})'
    for c in range(1, 10):
        if ws_sc.cell(row=1, column=c).value == prev_sc_label:
            ps = ws_sc.cell(row=2, column=c).value or 0
            res_stats['prev_avg_score'] = ps
            if ps: res_stats['score_change'] = (res_stats['avg_score'] - ps)/ps*100

    # 4. Time
    ws_ti = wb['gameplay_report(time) ']
    for c in range(1, 10):
        v = ws_ti.cell(row=1, column=c).value
        if v == f'Avg ({month_str.capitalize()})': 
            res_stats['avg_time'] = ws_ti.cell(row=2, column=c).value or 0
    prev_ti_label = f'Avg ({m_prev.capitalize()})'
    for c in range(1, 10):
        if ws_ti.cell(row=1, column=c).value == prev_ti_label:
            pt = ws_ti.cell(row=2, column=c).value or 0
            res_stats['prev_avg_time'] = pt
            if pt: res_stats['time_change'] = (res_stats['avg_time'] - pt)/pt*100

    # Build JSON
    output = {
        "report_title": f"{target_full} Report",
        "pages": [
            {
                "page_number": 3,
                "sections": {
                    "User Funnel": {
                        "metrics": {
                            "Total click": click,
                            "Register": reg,
                            "Player": play,
                            "Conversion rate": f"{conv_reg:.1f}%",
                            "Drop off": f"{drop_off:.1f}%"
                        }
                    },
                    "User Engagement": {
                        "comparison": {
                            "previous_month": prev_full,
                            "current_month": target_full
                        },
                        "metrics": {
                            "Daily Active Users (Avg.)": {
                                prev_full: round(res_stats['prev_dau'], 1),
                                target_full: round(res_stats['dau'], 1)
                            },
                            "Monthly Active Users": {
                                prev_full: res_stats['prev_mau'],
                                target_full: res_stats['mau']
                            },
                            "User Stickiness": {
                                prev_full: f"{res_stats['prev_stickiness'] * 100:.2f}%" if res_stats['prev_stickiness'] < 1 else f"{res_stats['prev_stickiness']:.2f}%",
                                target_full: f"{res_stats['stickiness'] * 100:.2f}%" if res_stats['stickiness'] < 1 else f"{res_stats['stickiness']:.2f}%"
                            }
                        }
                    }
                }
            },
            {
                "page_number": 4,
                "title": "Game Performance (Score)",
                "metrics": {
                    "AVG Score per Day": {
                        prev_full: int(res_stats.get('prev_avg_score', 0)),
                        target_full: int(res_stats['avg_score']),
                        "difference": f"{res_stats['score_change']:.1f}%"
                    }
                }
            },
            {
                "page_number": 5,
                "title": "Game Performance (Time)",
                "metrics": {
                    "AVG Time per Day": {
                        prev_full: f"{res_stats.get('prev_avg_time', 0):.0f} minute",
                        target_full: f"{res_stats['avg_time']:.0f} minute",
                        "difference": f"{res_stats['time_change']:.1f}%"
                    }
                }
            }
        ]
    }

    output_filename = f"value_data_{month_str}.json"
    with open(output_filename, 'w') as f:
        json.dump(output, f, indent=4)
        
    print(f"Data extracted and saved to {output_filename}")
    print(json.dumps(output, indent=2))

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python extract_data.py <excel_file> <target_month>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    target_month = sys.argv[2]
    
    extract_data(excel_file, target_month)