import openpyxl

def inspect_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_of_interest = ['User_funnel', 'User Engagement', 'gameplay_report(score) ', 'gameplay_report(time) ']
        
        for sheet_name in sheets_of_interest:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n--- Sheet: '{sheet_name}' ---")
                
                # Check for charts
                if ws._charts:
                    for i, chart in enumerate(ws._charts):
                        print(f"Chart {i+1}: Type={type(chart).__name__}")
                        try:
                            for j, series in enumerate(chart.series):
                                print(f"  Series {j+1}:")
                                if hasattr(series, 'title') and series.title and hasattr(series.title, 'tx') and hasattr(series.title.tx, 'strRef'):
                                    print(f"    Title Ref: {series.title.tx.strRef.f if hasattr(series.title.tx.strRef, 'f') else 'Unknown'}")
                                if hasattr(series, 'val') and hasattr(series.val, 'numRef'):
                                    print(f"    Val Ref: {series.val.numRef.f}")
                                if hasattr(series, 'cat') and hasattr(series.cat, 'strRef'):
                                    print(f"    Cat Ref: {series.cat.strRef.f}")
                                elif hasattr(series, 'cat') and hasattr(series.cat, 'numRef'):
                                    print(f"    Cat Ref (num): {series.cat.numRef.f}")
                        except Exception as e:
                            print(f"  Could not read series details: {e}")
                else:
                    print("No charts found.")
                    
                # Look for months string in the sheet to find the rows
                print("Searching for month markers (Jan, Feb, Mar)...")
                for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
                    for cell in row:
                        val = cell.value
                        if isinstance(val, str) and any(m in val.lower() for m in ['jan', 'feb', 'mar']):
                            print(f"Found '{val}' at {cell.coordinate}")
            else:
                print(f"\nSheet '{sheet_name}' not found.")
                
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    inspect_excel('Aji_game copy.xlsx')
